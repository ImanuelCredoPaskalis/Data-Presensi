"""
Export Utility - Export data presensi Mahasiswa ke Excel (.xlsx) dan CSV
Mendukung Multi-Sesi Kelas, Rincian Per Sesi Kelas, dan Akumulasi Durasi Total Kelas
"""
import csv
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import load_config
from database import calculate_time_diff_hours

def build_excel_workbook(records, start_date=None, end_date=None):
    """
    Membuat objek openpyxl.Workbook dari data presensi mahasiswa.
    """
    config = load_config()
    company_name = config.get("company_name", "Lab Micro Teaching FisMat")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Presensi Mahasiswa"

    # Style definitions
    font_title = Font(name="Segoe UI", size=15, bold=True, color="1E3A8A")
    font_sub = Font(name="Segoe UI", size=10, italic=True, color="475569")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)

    fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    fill_status_tepat = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_status_terlambat = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fill_status_tugas = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    fill_status_kelas = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Title Block
    ws.merge_cells("A1:M1")
    ws["A1"] = f"LAPORAN REKAPITULASI PRESENSI MAHASISWA - {company_name.upper()}"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    periode_str = "Semua Riwayat"
    if start_date and end_date:
        periode_str = f"Periode: {start_date} s/d {end_date}"
    elif start_date:
        periode_str = f"Mulai Tanggal: {start_date}"

    ws.merge_cells("A2:M2")
    ws["A2"] = f"{periode_str} | Dicetak pada: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    ws["A2"].font = font_sub
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Header Columns (Row 4)
    headers = [
        "No",
        "Tanggal",
        "Nama Mahasiswa",
        "Program Studi",
        "Jam Masuk",
        "Sesi Kelas",
        "Rincian Jam Kelas",
        "Total Durasi Kelas",
        "Tugas Luar",
        "Kembali Tugas",
        "Jam Keluar",
        "Durasi Total",
        "Status & Keterangan"
    ]

    header_row = 4
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header_title)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_cell
    
    ws.row_dimensions[header_row].height = 28

    # Data Rows
    current_row = 5
    for idx, r in enumerate(records, 1):
        jam_masuk = r.get("jam_masuk") or "-"
        
        # Multi-sesi kelas info
        total_sesi = r.get("total_sesi_kelas", 0)
        sesi_str = f"{total_sesi} Sesi" if total_sesi > 0 else "-"
        rincian_kelas = r.get("ringkasan_kelas") or "-"
        durasi_total_kelas = r.get("durasi_total_kelas") or "-"
        
        jam_tugas = r.get("jam_bertugas_keluar") or "-"
        jam_kembali = r.get("jam_kembali") or "-"
        jam_keluar = r.get("jam_keluar") or "-"
        durasi_kerja = calculate_time_diff_hours(r.get("jam_masuk"), r.get("jam_keluar"))
        
        status_ket = r.get("status") or "Hadir"
        keterangan_tambahan = []
        if r.get("keterangan_tugas"):
            keterangan_tambahan.append(f"Tugas: {r.get('keterangan_tugas')}")
        
        if keterangan_tambahan:
            status_ket += f" ({', '.join(keterangan_tambahan)})"

        row_values = [
            idx,
            r.get("tanggal", "-"),
            r.get("nama", "-"),
            r.get("departemen", "-"),
            jam_masuk,
            sesi_str,
            rincian_kelas,
            durasi_total_kelas,
            jam_tugas,
            jam_kembali,
            jam_keluar,
            durasi_kerja,
            status_ket
        ]

        row_fill = fill_even if idx % 2 == 0 else fill_odd

        for col_num, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_num, value=val)
            cell.font = font_data
            cell.fill = row_fill
            cell.border = border_cell

            # Alignment logic
            if col_num in [1, 2, 5, 6, 8, 9, 10, 11, 12]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Status highlighting
            if col_num == 13:
                if "Terlambat" in str(val):
                    cell.fill = fill_status_terlambat
                elif "Tepat Waktu" in str(val):
                    cell.fill = fill_status_tepat
                elif "Kelas" in str(val):
                    cell.fill = fill_status_kelas
                elif "Tugas Luar" in str(val):
                    cell.fill = fill_status_tugas

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # Auto-adjust column widths
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 11)

    return wb

def export_to_excel(records, file_path, start_date=None, end_date=None):
    """
    Ekspor list data presensi mahasiswa ke file Excel dengan format rapi & profesional.
    """
    wb = build_excel_workbook(records, start_date, end_date)
    wb.save(file_path)
    return True

def get_excel_bytes(records, start_date=None, end_date=None):
    """
    Menghasilkan data binary Excel (.xlsx) di memory untuk download via web browser.
    """
    import io
    wb = build_excel_workbook(records, start_date, end_date)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def export_to_csv(records, file_path):
    """
    Ekspor list data presensi mahasiswa ke file CSV.
    """
    with open(file_path, mode="w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "No", "Tanggal", "Nama Mahasiswa", "Program Studi", "Status Mahasiswa",
            "Jam Masuk", "Total Sesi Kelas", "Rincian Jam Kelas", "Total Durasi Kelas",
            "Jam Tugas Keluar", "Jam Kembali Tugas", "Jam Keluar",
            "Durasi Total", "Status", "Keterangan Tugas"
        ])

        for idx, r in enumerate(records, 1):
            total_sesi = r.get("total_sesi_kelas", 0)
            rincian_kelas = r.get("ringkasan_kelas") or ""
            durasi_total_kelas = r.get("durasi_total_kelas") or ""
            durasi_kerja = calculate_time_diff_hours(r.get("jam_masuk"), r.get("jam_keluar"))
            
            writer.writerow([
                idx,
                r.get("tanggal", ""),
                r.get("nama", ""),
                r.get("departemen", ""),
                r.get("jabatan", "Mahasiswa"),
                r.get("jam_masuk", ""),
                total_sesi,
                rincian_kelas,
                durasi_total_kelas,
                r.get("jam_bertugas_keluar", ""),
                r.get("jam_kembali", ""),
                r.get("jam_keluar", ""),
                durasi_kerja,
                r.get("status", ""),
                r.get("keterangan_tugas", "")
            ])
    return True

def get_csv_bytes(records):
    """
    Menghasilkan data binary CSV di memory untuk download via web browser.
    """
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "No", "Tanggal", "Nama Mahasiswa", "Program Studi", "Status Mahasiswa",
        "Jam Masuk", "Total Sesi Kelas", "Rincian Jam Kelas", "Total Durasi Kelas",
        "Jam Tugas Keluar", "Jam Kembali Tugas", "Jam Keluar",
        "Durasi Total", "Status", "Keterangan Tugas"
    ])

    for idx, r in enumerate(records, 1):
        total_sesi = r.get("total_sesi_kelas", 0)
        rincian_kelas = r.get("ringkasan_kelas") or ""
        durasi_total_kelas = r.get("durasi_total_kelas") or ""
        durasi_kerja = calculate_time_diff_hours(r.get("jam_masuk"), r.get("jam_keluar"))
        
        writer.writerow([
            idx,
            r.get("tanggal", ""),
            r.get("nama", ""),
            r.get("departemen", ""),
            r.get("jabatan", "Mahasiswa"),
            r.get("jam_masuk", ""),
            total_sesi,
            rincian_kelas,
            durasi_total_kelas,
            r.get("jam_bertugas_keluar", ""),
            r.get("jam_kembali", ""),
            r.get("jam_keluar", ""),
            durasi_kerja,
            r.get("status", ""),
            r.get("keterangan_tugas", "")
        ])
    return output.getvalue().encode("utf-8-sig")
